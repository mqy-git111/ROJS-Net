import numpy
import numpy as np
import SimpleITK as sitk
import openpyxl
import torch.nn.functional as F
import config
import torch
import os
import csv
from utils import common
from medpy import metric
import json



def load_file_name_list(file_path):
    file_name_list = []
    with open(file_path, 'r') as file_to_read:
        while True:
            lines = file_to_read.readline().strip()  # 整行读取数据
            if not lines:
                break
            file_name_list.append(lines.split())
    return file_name_list
def getdice(logits,targets):
    inter = torch.sum(logits * targets)
    union = torch.sum(logits) + torch.sum(targets)
    dice = (2. * inter + 1e-8) / (union + 1e-8)
    return dice
def getasd(logits,targets):
    asd = 0
    for i in range(64):
        if np.sum(logits[0][i].numpy())== 0 and np.sum(targets[0][i].numpy()) == 0:
            asd_i = 0
        if np.sum(logits[0][i].numpy()) == 0 or np.sum(targets[0][i].numpy()) == 0:
            asd_i = 1
        else:
            asd_i = metric.asd(logits[0][i].numpy(),targets[0][i].numpy())
        asd = asd + asd_i
    asd = asd / 64
    return asd
def gethd(logits,targets):
    hd = 0
    for i in range(64):
        if np.sum(logits[0][i].numpy())== 0 and np.sum(targets[0][i].numpy()) == 0:
            hd_i = 0
        if np.sum(logits[0][i].numpy()) == 0 or np.sum(targets[0][i].numpy()) == 0:
            hd_i = 1
        else:
            hd_i = metric.hd95(logits[0][i].numpy(),targets[0][i].numpy())
        hd = hd + hd_i
    hd = hd / 64
    return hd

def get_hd(predLabel_array,labels,name,spacing):
    # 计算dice
    preorgan_fdata = predLabel_array[0].squeeze()
    pretumor_fdata = predLabel_array[1].squeeze()
    preorgan_fdata = torch.FloatTensor(preorgan_fdata.astype(float)).unsqueeze(0)
    pretumor_fdata = torch.FloatTensor(pretumor_fdata.astype(float)).unsqueeze(0)
    preorgan_labels = common.to_one_hot_3d(preorgan_fdata.long(), 4).numpy()
    pretumor_labels = common.to_one_hot_3d(pretumor_fdata.long(), 2).numpy()
    labelorgan_fdata = labels[0].squeeze()
    labeltumor_fdata = labels[1].squeeze()
    labelorgan_fdata = torch.FloatTensor(labelorgan_fdata.astype(float)).unsqueeze(0)
    labeltumor_fdata = torch.FloatTensor(labeltumor_fdata.astype(float)).unsqueeze(0)
    organ_label = common.to_one_hot_3d(labelorgan_fdata.long(), 4).numpy()
    tumor_label = common.to_one_hot_3d(labeltumor_fdata.long(), 2).numpy()
    temp_hd = [name]
    hds = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
    for i in range(1, 4):
        hd = gethd(torch.FloatTensor(organ_label[:, i, :, :, :]),
                       torch.FloatTensor(preorgan_labels[:, i, :, :, :])) * spacing
        hd = np.array(hd) * 3  #3是z轴spacing
        temp_hd.append(hd)
        hds[i] = hds[i] + hd
    hd = gethd(torch.FloatTensor(tumor_label[:, 1, :, :, :]),
                   torch.FloatTensor(pretumor_labels[:, 1, :, :, :])) * spacing
    hd = np.array(hd) * 3
    hds[4] = hds[4] + hd
    temp_hd.append(hd*3)
    # print(temp_hd)
    return temp_hd


def get_asd(predLabel_array,labels,name,spacing):
    # 计算dice
    preorgan_fdata = predLabel_array[0].squeeze()
    pretumor_fdata = predLabel_array[1].squeeze()
    preorgan_fdata = torch.FloatTensor(preorgan_fdata.astype(float)).unsqueeze(0)
    pretumor_fdata = torch.FloatTensor(pretumor_fdata.astype(float)).unsqueeze(0)
    preorgan_labels = common.to_one_hot_3d(preorgan_fdata.long(), 4).numpy()
    pretumor_labels = common.to_one_hot_3d(pretumor_fdata.long(), 2).numpy()
    labelorgan_fdata = labels[0].squeeze()
    labeltumor_fdata = labels[1].squeeze()
    labelorgan_fdata = torch.FloatTensor(labelorgan_fdata.astype(float)).unsqueeze(0)
    labeltumor_fdata = torch.FloatTensor(labeltumor_fdata.astype(float)).unsqueeze(0)
    organ_label = common.to_one_hot_3d(labelorgan_fdata.long(), 4).numpy()
    tumor_label = common.to_one_hot_3d(labeltumor_fdata.long(), 2).numpy()
    temp_asd = [name]
    asds = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
    for i in range(1, 4):
        asd = getasd(torch.FloatTensor(organ_label[:, i, :, :, :]),
                       torch.FloatTensor(preorgan_labels[:, i, :, :, :])) * spacing
        asd = np.array(asd) * 3  #3是z轴spacing
        temp_asd.append(asd)
        asds[i] = asds[i] + asd
    asd = getasd(torch.FloatTensor(tumor_label[:, 1, :, :, :]),
                   torch.FloatTensor(pretumor_labels[:, 1, :, :, :])) * spacing
    asd = np.array(asd) * 3
    asds[4] = asds[4] + asd
    temp_asd.append(asd*3)
    # print(temp_asd)
    return temp_asd
def get_dice(ground,pred,num,temp_dice,count):
    # 计算dice
    ground = ground.squeeze()
    pred = pred.squeeze()
    ground = torch.FloatTensor(ground.astype(float)).unsqueeze(0)
    pred = torch.FloatTensor(pred.astype(float)).unsqueeze(0)
    ground = common.to_one_hot_3d(ground.long(), num).numpy()
    pred = common.to_one_hot_3d(pred.long(), num).numpy()
    for i in range(1, num):
        dice = getdice(torch.FloatTensor(ground[:, i, :, :, :]),
                       torch.FloatTensor(pred[:, i, :, :, :])) * 100
        dice = np.array(dice).item()
        if ground[:, i, :, :, :].sum() == 0:
            temp_dice.append(0.0)
        else:
            temp_dice.append(dice)
            count[i] = count[i] + 1
    return temp_dice,count

save_path = "/home/user_gou/whole_body_seg/Data/nnUNet_raw/Dataset104_abdomen"
labelsTs = save_path + "/test/labelsTs"
predict = save_path + "/test/predict"
json_file = save_path + "/dataset.json"
f_path = save_path + "/dice_info.xlsx"
filename_list = os.listdir(labelsTs)
count = [0 for i in range(0,60)]
dice_all = []
workbook = openpyxl.Workbook()
sheet = workbook.active
#读取json文件获得器官名称
f = open(json_file, 'r')
content = f.read()
a = json.loads(content)
labels = a["labels"]
f.close()
data =[]
col_num = 1
# key使用lambda匿名函数按键进行排序
a2 = sorted(labels.items(),key = lambda x:x[1])
data = [i[0] for i in a2]
labels_num = len(data)
for row, value in enumerate(data, start=1):
    sheet.cell(row=row, column=col_num, value=value)
col_num = col_num + 1
for singleImg in filename_list:
    singleImg = "AbdomenCT-1K-Subtask1_0011.nii.gz"
    label = sitk.ReadImage(os.path.join(labelsTs, singleImg))
    label_fdata = sitk.GetArrayFromImage(label)
    pred = sitk.ReadImage(os.path.join(predict, singleImg))
    pred_fdata = sitk.GetArrayFromImage(pred)
    dice,count = get_dice(label_fdata,pred_fdata,labels_num,[singleImg],count)
    for row, value in enumerate(data, start=1):
        sheet.cell(row=row, column=col_num, value=value)
    col_num = col_num + 1
    dice_all.append(dice[1:])
    break
    # hd = get_hd([pre_organ,pre_tumor],[organ_fdata,tumor_fdata],name,spacing)
    # asd = get_asd([pre_organ,pre_tumor],[organ_fdata,tumor_fdata],name,spacing)

#     hd_all.append(hd[1:])
#     asd_all.append(asd[1:])
#     num = num + 1
# # dice_avg
dice_avg = np.sum(dice_all,axis=0)
count = count[1:]
res_dice = [dice_avg[i]/count[i] for i in range(0, 59)]
for row, value in enumerate(res_dice, start=2):
    sheet.cell(row=row, column=col_num, value=value)
amap = dict()
for i in range(0,59):
    amap[data[i + 1]] = res_dice[i]
workbook.save(f_path)
print(amap)

# organ_avg_dice = np.average(dice_avg[:-1])
# all_dice_avg = np.average(dice_avg)
# dice_avg = list(dice_avg)
# dice_avg.append("avg")
# dice_avg.append(organ_avg_dice)
# dice_avg.append(all_dice_avg)
# dice_std = np.std(dice_all,axis=0)
# organ_std_dice = np.average(dice_std[:-1])
# all_dice_std = np.average(dice_std)
# dice_std = list(dice_std)
# dice_std.append("avg")
# dice_std.append(organ_std_dice)
# dice_std.append(all_dice_std)
# print(["dice_avg"] + dice_avg)
# print(["dice_std"] + dice_std)
# writer.writerow(["dice_avg"] + dice_avg)
# writer.writerow(["dice_std"] + dice_std)
#
# # hd_avg
# hd_avg = np.average(hd_all,axis=0)
# organ_avg_hd = np.average(hd_avg[:-1])
# all_hd_avg = np.average(hd_avg)
# hd_avg = list(hd_avg)
# hd_avg.append("avg")
# hd_avg.append(organ_avg_hd)
# hd_avg.append(all_hd_avg)
# hd_std = np.std(hd_all,axis=0)
# organ_std_hd = np.average(hd_std[:-1])
# all_hd_std = np.average(hd_std)
# hd_std = list(hd_std)
# hd_std.append("avg")
# hd_std.append(organ_std_hd)
# hd_std.append(all_hd_std)
# print(["hd_avg"] + hd_avg)
# print(["hd_std"] + hd_std)
# writer.writerow(["hd_avg"] + hd_avg)
# writer.writerow(["hd_std"] + hd_std)
#
#
# #asd_avg
# asd_avg = np.average(asd_all,axis=0)
# organ_avg_asd = np.average(asd_avg[:-1])
# all_asd_avg = np.average(asd_avg)
# asd_avg = list(asd_avg)
# asd_avg.append("avg")
# asd_avg.append(organ_avg_asd)
# asd_avg.append(all_asd_avg)
# asd_std = np.std(asd_all,axis=0)
# organ_std_asd = np.average(asd_std[:-1])
# all_asd_std = np.average(asd_std)
# asd_std = list(asd_std)
# asd_std.append("avg")
# asd_std.append(organ_std_asd)
# asd_std.append(all_asd_std)
# print(["asd_avg"] + asd_avg)
# print(["asd_std"] + asd_std)
# writer.writerow(["asd_avg"] + asd_avg)
# writer.writerow(["asd_std"] + asd_std)
#
# f.close()
# print()
